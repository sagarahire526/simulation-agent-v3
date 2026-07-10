"""
scripts/scenario_consistency_test.py

Automated consistency test for the DETERMINISTIC scenario nodes. This complements
scripts/run_consistency_sheet.py (a manual, end-to-end 4-run eyeball review of the full
SSE pipeline). Instead, this harness ISOLATES the two layers that matter for consistency
and asserts them automatically:

  1. DETERMINISM  — run_scenario(scn, filter, group_by) has NO LLM in it (pure SQL + fixed
                    math). So the SAME filter run N times MUST be byte-identical. Only
                    CURRENT_DATE varies, so "identical" means same-day, same-filter.
  2. VARIATION    — different filters (e.g. region SOUTH vs WEST) MUST yield different
                    output — catches a scenario that ignores a filter and returns a constant.
  3. EXTRACTION   — the ONE non-deterministic step is extract_params_by_schema (a constrained
                    LLM call). Same canonical question run N times MUST resolve to stable params.

Layer 1/2 hit the real read-only Postgres via the scenario's own node functions; layer 3
makes N small LLM calls per scenario.

The filters for layers 1/2 are generated FROM each scenario's own scn_param_schema (fetched
from Neo4j), so the harness stays correct as schemas evolve.

Output: a workbook (xlsx if openpyxl is present, else CSV files) with a Determinism sheet,
a Variation sheet, an Extraction sheet, and a Summary.

Usage:
  venv/bin/python scripts/scenario_consistency_test.py
  venv/bin/python scripts/scenario_consistency_test.py --runs 3 --out scenario_consistency.xlsx
  venv/bin/python scripts/scenario_consistency_test.py --scenario scn-008-crew-shortage-plan
  venv/bin/python scripts/scenario_consistency_test.py --no-extraction   # skip the LLM layer
"""
from __future__ import annotations

import argparse
import json
import os
import sys
from datetime import date

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import services.sandbox_service as sandbox_svc  # noqa: E402
from tools.neo4j_tool import Neo4jTool  # noqa: E402

# Program used to scope every test (the extractor injects smp_name/project_type from the
# user's selection; here we fix it so results are comparable).
TEST_SMP = "NTM"
TEST_PROJECT_TYPE = "NTM"
REGIONS_FOR_VARIATION = ["SOUTH", "WEST"]


def _canon(obj) -> str:
    """Canonical, order-independent JSON string for equality comparison."""
    return json.dumps(obj, sort_keys=True, default=str)


def _fetch_scenarios(only: str | None) -> list[dict]:
    t = Neo4jTool()
    try:
        r = t.run_cypher(
            "MATCH (n:BKGNode {entity_type:'scenario'}) "
            "RETURN n.node_id AS node_id, n.label AS label, "
            "coalesce(n.scn_canonical_question,'') AS q, "
            "coalesce(n.scn_param_schema,'') AS schema ORDER BY n.node_id"
        )
    finally:
        t.close()
    rows = r.get("records", []) if r.get("status") == "success" else []
    out = []
    for rec in rows:
        if only and rec["node_id"] != only:
            continue
        schema = None
        try:
            schema = json.loads(rec["schema"]) if rec["schema"].strip() else None
        except json.JSONDecodeError:
            schema = None
        out.append({"node_id": rec["node_id"], "label": rec["label"],
                    "question": rec["q"], "schema": schema})
    return out


def _sample_value(field: dict):
    """A representative non-empty value for a schema field (for filter generation)."""
    name = field.get("name", "")
    typ = (field.get("type") or "string").lower()
    default = field.get("default")
    if default not in (None, 0, [], ""):
        return default
    if typ == "int":
        if "target" in name or "rate" in name:
            return 1000
        if "delay" in name:
            return 1
        return 6 if "horizon" in name else 2
    if typ == "number":
        if "pct" in name or "shortage" in name or "increase" in name or "fasttrack" in name:
            return 15
        return 20
    if typ == "enum" and field.get("allowed"):
        return field["allowed"][0]
    if typ == "enum_list" and field.get("allowed"):
        return field["allowed"][:2]
    return None


def _build_filter(scn: dict, rgn_region: str | None):
    """Build a concrete (filter, group_by) for a scenario from its schema, scoping to
    the given region. Legacy scn-001 (no 'fields' schema) is handled specially."""
    schema = scn["schema"] or {}
    fields = schema.get("fields")
    if not fields:
        # Legacy SCOP scenario: backward look-back window + program + group_by.
        today = date.today()
        start = date(today.year, max(1, today.month - 2), 1).isoformat()
        filt = {"start_date": start, "end_date": today.isoformat(),
                "smp_name": TEST_SMP}
        if rgn_region:
            filt["rgn_region"] = rgn_region
        return filt, "construction_gc"

    gbf = schema.get("group_by_field")
    filt = {}
    group_by = None
    for fld in fields:
        nm = fld.get("name")
        if not nm:
            continue
        if nm == "rgn_region":
            if rgn_region:
                filt["rgn_region"] = rgn_region
            continue
        val = _sample_value(fld)
        if val is None:
            continue
        if gbf and nm == gbf:
            group_by = val
        else:
            filt[nm] = val
    filt["smp_name"] = TEST_SMP
    filt["project_type"] = TEST_PROJECT_TYPE
    return filt, group_by


def _run(scn_id: str, filt: dict, group_by, timeout: int):
    out = sandbox_svc.run_bkg_node("scenario", scn_id, filters=filt,
                                   group_by=group_by, timeout_seconds=timeout)
    if out.get("status") != "success":
        return None, str(out.get("error"))[:300]
    return out.get("result"), None


def _summary_metric(result) -> str:
    """A short human-readable metric string from a scenario result (for the sheet)."""
    if not isinstance(result, dict):
        return str(result)[:80]
    keys = ("target_sites", "sites_at_risk", "revised_weekly_run_rate", "pending_count",
            "total_crews", "po_gate", "cohorts", "predictions")
    bits = []
    for k in keys:
        if k in result:
            v = result[k]
            if isinstance(v, (list, dict)):
                v = f"<{len(v)}>"
            bits.append(f"{k}={v}")
    return ", ".join(bits)[:120] or ", ".join(list(result.keys())[:6])


def run_determinism(scenarios, runs, timeout):
    """Layer 1: same filter N times -> byte-identical. Returns rows + variation cache."""
    rows = []
    var_cache = {}  # node_id -> {region: canonical_json}
    for scn in scenarios:
        for rgn in REGIONS_FOR_VARIATION:
            filt, gb = _build_filter(scn, rgn)
            canon = None
            identical = True
            err = None
            first_result = None
            for i in range(runs):
                res, e = _run(scn["node_id"], filt, gb, timeout)
                if e:
                    err = e
                    identical = False
                    break
                if first_result is None:
                    first_result = res
                c = _canon(res)
                if canon is None:
                    canon = c
                elif c != canon:
                    identical = False
            var_cache.setdefault(scn["node_id"], {})[rgn] = canon
            rows.append({
                "scenario": scn["node_id"], "region": rgn, "group_by": gb,
                "runs": runs,
                "deterministic": "ERROR" if err else ("PASS" if identical else "FAIL"),
                "metric": _summary_metric(first_result) if first_result is not None else "",
                "error": err or "",
                "filter": _canon(filt),
            })
            status = "ERR " if err else ("PASS" if identical else "FAIL")
            print(f"  [det] {scn['node_id']:38} {rgn:7} {status}  {rows[-1]['metric'][:60]}")
    return rows, var_cache


def run_variation(scenarios, var_cache):
    """Layer 2: region SOUTH vs WEST must differ (unless both legitimately empty)."""
    rows = []
    for scn in scenarios:
        cache = var_cache.get(scn["node_id"], {})
        a = cache.get(REGIONS_FOR_VARIATION[0])
        b = cache.get(REGIONS_FOR_VARIATION[1])
        if a is None or b is None:
            verdict, note = "SKIP", "a region errored"
        elif a == b:
            # Same output for two regions — suspicious unless both are trivially empty.
            verdict = "REVIEW"
            note = "SOUTH == WEST output (check the region filter is applied / both empty)"
        else:
            verdict, note = "PASS", "regions differ as expected"
        rows.append({"scenario": scn["node_id"],
                     "varied": f"rgn_region {REGIONS_FOR_VARIATION[0]} vs {REGIONS_FOR_VARIATION[1]}",
                     "verdict": verdict, "note": note})
        print(f"  [var] {scn['node_id']:38} {verdict}  {note}")
    return rows


def run_extraction(scenarios, runs):
    """Layer 3: same canonical question -> stable resolved params across N runs."""
    from services.scenario_params import extract_params_by_schema, extract_scenario_params
    rows = []
    for scn in scenarios:
        q = scn["question"]
        if not q:
            continue
        resolved = []
        err = None
        for _ in range(runs):
            try:
                if scn["schema"] and scn["schema"].get("fields"):
                    p = extract_params_by_schema(q, scn["schema"], project_type=TEST_PROJECT_TYPE)
                else:
                    p = extract_scenario_params(q, project_type=TEST_PROJECT_TYPE)
                resolved.append(_canon(p.get("resolved", p)))
            except Exception as e:  # noqa: BLE001
                err = str(e)[:200]
                break
        stable = bool(resolved) and all(r == resolved[0] for r in resolved) and not err
        rows.append({"scenario": scn["node_id"], "runs": runs,
                     "stable": "ERROR" if err else ("PASS" if stable else "FAIL"),
                     "params": resolved[0] if resolved else "", "error": err or ""})
        print(f"  [ext] {scn['node_id']:38} {'ERR' if err else ('PASS' if stable else 'FAIL')}")
    return rows


def write_output(out_path, det_rows, var_rows, ext_rows):
    try:
        import openpyxl
        wb = openpyxl.Workbook()
        def sheet(title, headers, rows):
            ws = wb.create_sheet(title) if wb.sheetnames != ["Sheet"] or title != "Determinism" else wb.active
            ws.title = title
            ws.append(headers)
            for r in rows:
                ws.append([r.get(h) for h in headers])
        sheet("Determinism", ["scenario", "region", "group_by", "runs", "deterministic", "metric", "error", "filter"], det_rows)
        sheet("Variation", ["scenario", "varied", "verdict", "note"], var_rows)
        if ext_rows is not None:
            sheet("Extraction", ["scenario", "runs", "stable", "params", "error"], ext_rows)
        # summary
        ws = wb.create_sheet("Summary")
        det_pass = sum(1 for r in det_rows if r["deterministic"] == "PASS")
        ws.append(["Determinism PASS", det_pass, "of", len(det_rows)])
        ws.append(["Determinism FAIL", sum(1 for r in det_rows if r["deterministic"] == "FAIL")])
        ws.append(["Determinism ERROR", sum(1 for r in det_rows if r["deterministic"] == "ERROR")])
        if ext_rows is not None:
            ws.append(["Extraction PASS", sum(1 for r in ext_rows if r["stable"] == "PASS"), "of", len(ext_rows)])
        wb.save(out_path)
        return out_path
    except ImportError:
        import csv
        base = os.path.splitext(out_path)[0]
        for name, headers, rows in [("determinism", ["scenario","region","group_by","runs","deterministic","metric","error","filter"], det_rows),
                                    ("variation", ["scenario","varied","verdict","note"], var_rows),
                                    ("extraction", ["scenario","runs","stable","params","error"], ext_rows or [])]:
            with open(f"{base}_{name}.csv", "w", newline="") as fh:
                w = csv.DictWriter(fh, fieldnames=headers)
                w.writeheader()
                w.writerows(rows)
        return f"{base}_*.csv"


def main():
    ap = argparse.ArgumentParser(description="Automated consistency test for scenario nodes.")
    ap.add_argument("--runs", type=int, default=2, help="runs per case for determinism/extraction (default 2)")
    ap.add_argument("--timeout", type=int, default=240, help="sandbox timeout seconds per run")
    ap.add_argument("--scenario", default=None, help="test only this node_id")
    ap.add_argument("--no-extraction", action="store_true", help="skip the LLM extraction layer")
    ap.add_argument("--out", default="scenario_consistency.xlsx")
    a = ap.parse_args()

    scenarios = _fetch_scenarios(a.scenario)
    if not scenarios:
        print("No scenario nodes found."); sys.exit(1)
    print(f"Testing {len(scenarios)} scenario node(s), {a.runs} run(s) each.\n")

    print("== Layer 1: determinism (same filter -> byte-identical) ==")
    det_rows, var_cache = run_determinism(scenarios, a.runs, a.timeout)
    print("\n== Layer 2: variation (region SOUTH vs WEST -> different) ==")
    var_rows = run_variation(scenarios, var_cache)
    ext_rows = None
    if not a.no_extraction:
        print("\n== Layer 3: extraction stability (same question -> stable params) ==")
        ext_rows = run_extraction(scenarios, a.runs)

    path = write_output(a.out, det_rows, var_rows, ext_rows)
    dp = sum(1 for r in det_rows if r["deterministic"] == "PASS")
    df = sum(1 for r in det_rows if r["deterministic"] == "FAIL")
    de = sum(1 for r in det_rows if r["deterministic"] == "ERROR")
    print(f"\nDeterminism: {dp} PASS / {df} FAIL / {de} ERROR  (of {len(det_rows)} cases)")
    if ext_rows is not None:
        print(f"Extraction : {sum(1 for r in ext_rows if r['stable']=='PASS')} PASS of {len(ext_rows)}")
    print(f"Wrote: {path}")


if __name__ == "__main__":
    main()
