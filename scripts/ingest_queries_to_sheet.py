"""
Run a batch of simulation queries in parallel against the SSE streaming endpoint
and dump the results into an .xlsx sheet.

Columns written (in order):
    ID  |  Simulation Query  |  Planner Steps  |  Response Format  |
    Data Correctness  |  KPI Used  |  Algorithm

Parallelism: `--concurrency N` SSE streams run at any time (default 5).
If you pass 10 queries with concurrency 5, the pool keeps 5 in-flight — as one finishes
the next is picked up — so effectively a rolling "batch of 5".

Endpoint used:
    GET  /api/v1/simulate/stream          (query, user_id, project_type, thread_id as query params)
    POST /api/v1/simulate/stream/resume   (auto-called with "Accept stated assumptions" on HITL)

Input file format (queries.txt):
    One query per line, query and project_type as quoted strings. Blank lines
    and '# comment' lines are ignored. Example:

        "How many active GC sites are in Chicago?"           "NTM"
        "What is completion rate for AHLOB in Chicago?"      "AHLOB Modernization"

Output .xlsx behaviour:
    - if the file does not exist → create it with headers
    - if the file exists         → append rows, continuing the ID counter
      from the last row already in the sheet

Usage
-----
venv/bin/python scripts/ingest_queries_to_sheet.py \
    --base-url http://localhost:8000 \
    --queries queries.txt \
    --output  results.xlsx

"""
from __future__ import annotations

import argparse
import json
import shlex
import sys
import threading
import time
import uuid
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path
from typing import Any
from urllib.parse import urlencode

import openpyxl
import requests
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


# ── Configuration ────────────────────────────────────────────────────────────

# The target sheet is organised as triplets: <field>, <field (Binary)>, <field Feedback>.
# The script fills only the data cells; Binary / Feedback columns are left blank
# for manual reviewer entry.
HEADERS = [
    "ID",
    "Simulation Query",
    "Planer Steps",                           # (sic — matches the existing sheet header)
    "Planner Steps Correct? (T/F)",
    "Planner Steps Feedback If 'F'",
    "Response Format",
    "Response Format Correct? (T/F)",
    "Response Format Feedback If 'F'",
    "Data",
    "Data Correct? (T/F)",
    "Data Correctness Feedback If 'F'",
    "KPIs Used",
    "KPI Used Correct? (T/F)",
    "KPIs Used Feedback If 'F'",
    "Algorithm Used",
    "Algorithm Used Correct? (T/F)",
    "Algorithm Feedback If 'F'",
]

# 1-indexed column positions the script writes into
COL_ID            = 1
COL_QUERY         = 2
COL_PLANNER_STEPS = 3
COL_RESPONSE      = 6
COL_DATA_CORRECT  = 9
COL_KPI_USED      = 12
COL_ALGORITHM     = 15
DATA_COLUMNS = {COL_ID, COL_QUERY, COL_PLANNER_STEPS, COL_RESPONSE,
                COL_DATA_CORRECT, COL_KPI_USED, COL_ALGORITHM}

COLUMN_WIDTHS = {
    "ID": 6,
    "Simulation Query": 50,
    "Planer Steps": 60,
    "Planner Steps Correct? (T/F)": 12,
    "Planner Steps Feedback If 'F'": 30,
    "Response Format": 70,
    "Response Format Correct? (T/F)": 12,
    "Response Format Feedback If 'F'": 30,
    "Data": 70,
    "Data Correct? (T/F)": 12,
    "Data Correctness Feedback If 'F'": 30,
    "KPIs Used": 60,
    "KPI Used Correct? (T/F)": 12,
    "KPIs Used Feedback If 'F'": 30,
    "Algorithm Used": 60,
    "Algorithm Used Correct? (T/F)": 12,
    "Algorithm Feedback If 'F'": 30,
}

REQUEST_TIMEOUT_SECONDS = 600  # simulation can run for several minutes


# ── Response parsing ─────────────────────────────────────────────────────────

def _parse_tool_output(raw: Any) -> dict | list | str:
    """tool_output comes back as a JSON string — try to parse it, fall back to raw."""
    if isinstance(raw, (dict, list)):
        return raw
    if not isinstance(raw, str):
        return raw
    try:
        return json.loads(raw)
    except (ValueError, TypeError):
        return raw


def _iter_tool_calls(traces: dict | None):
    """Yield every tool_call dict across all planner steps in order."""
    if not traces:
        return
    for step in traces.get("steps", []) or []:
        for tc in step.get("tool_calls", []) or []:
            yield tc


def _build_node_name_map(traces: dict | None) -> dict[str, str]:
    """
    Walk every get_node tool call and return {node_id → human-readable name}.
    The simulation_service trims get_node outputs to (node_id, name, label,
    map_python_function) — so `name` is available there even though get_kpi
    traces have it stripped out.
    """
    mapping: dict[str, str] = {}
    for tc in _iter_tool_calls(traces):
        if tc.get("tool_name") != "get_node":
            continue
        parsed = _parse_tool_output(tc.get("tool_output"))
        if not isinstance(parsed, dict):
            continue
        node_id = parsed.get("node_id") or (tc.get("tool_input") or {}).get("node_id")
        name = parsed.get("name")
        if node_id and name:
            mapping[node_id] = name
    return mapping


_KPI_SEPARATOR = "-" * 60


def extract_kpi_used(traces: dict | None) -> str:
    """
    Build the 'KPI Used' cell.

    Output format — each KPI followed by the code that consumed it,
    long rule between entries:

        Active GC Sites
        <python / sql code>
        ------------------------------------------------------------
        Next KPI Name
        <its code>
        ------------------------------------------------------------

    Code is attributed to the most recent preceding get_kpi call.
    Any trailing run_python / run_sql_python with no preceding KPI is
    appended under an "(Unassociated code)" entry.
    """
    name_by_id = _build_node_name_map(traces)

    # Group code calls under the nearest preceding get_kpi call.
    # Each group is (kpi_label, [code strings]).
    groups: list[tuple[str, list[str]]] = []
    current: tuple[str, list[str]] | None = None

    def _kpi_label_from(tc: dict) -> str:
        tool_input = tc.get("tool_input") or {}
        tool_output = _parse_tool_output(tc.get("tool_output"))
        kpi_name = ""
        node_id = ""
        if isinstance(tool_output, dict):
            kpi_name = tool_output.get("kpi_name") or ""
            node_id = tool_output.get("kpi_kpi_id") or ""
        node_id = node_id or tool_input.get("node_id") or ""
        return kpi_name or name_by_id.get(node_id) or node_id or "unknown KPI"

    for tc in _iter_tool_calls(traces):
        tool_name = tc.get("tool_name", "")
        tool_input = tc.get("tool_input") or {}

        if tool_name == "get_kpi":
            if current is not None:
                groups.append(current)
            current = (_kpi_label_from(tc), [])

        elif tool_name in ("run_python", "run_sql_python"):
            # Keep only code that actually ran — failed SQL/Python calls are noise.
            if tc.get("status") != "success":
                continue
            code = tool_input.get("code")
            if not code:
                continue
            if current is None:
                current = ("(Unassociated code)", [])
            current[1].append(code)

    if current is not None:
        groups.append(current)

    if not groups:
        return ""

    # Dedupe KPI entries by (label, code) — an agent sometimes re-fetches the same KPI.
    seen: set[tuple[str, str]] = set()
    blocks: list[str] = []
    for label, codes in groups:
        joined_code = "\n\n".join(codes).strip()
        key = (label, joined_code)
        if key in seen:
            continue
        seen.add(key)
        block = label if not joined_code else f"{label}\n{joined_code}"
        blocks.append(block)

    return ("\n" + _KPI_SEPARATOR + "\n").join(blocks)


def extract_algorithm(api_response: dict | None, traces: dict | None) -> str:
    """
    Prefer the `execution_algorithm` narrative produced by the fast-tier LLM
    that runs in parallel with the response agent. If the server is older and
    didn't emit that field, fall back to concatenating kpi_business_logic from
    every get_kpi call — that's what this column used to hold.
    """
    if api_response:
        narrative = (api_response.get("execution_algorithm") or "").strip()
        if narrative:
            return narrative

    name_by_id = _build_node_name_map(traces)
    algos: list[str] = []
    seen: set[str] = set()

    for tc in _iter_tool_calls(traces):
        if tc.get("tool_name") != "get_kpi":
            continue
        parsed = _parse_tool_output(tc.get("tool_output"))
        if not isinstance(parsed, dict):
            continue
        logic = parsed.get("kpi_business_logic")
        if not logic or not isinstance(logic, str):
            continue
        logic = logic.strip()
        if not logic or logic in seen:
            continue
        seen.add(logic)
        node_id = parsed.get("kpi_kpi_id") or (tc.get("tool_input") or {}).get("node_id") or ""
        label = parsed.get("kpi_name") or name_by_id.get(node_id) or node_id
        algos.append(f"[{label}]\n{logic}" if label else logic)

    return "\n\n---\n\n".join(algos)


def format_planner_steps(steps: list[str] | None, traces: dict | None = None) -> str:
    """
    Render the Planner Steps cell. For simulation-path runs this is the
    decomposed plan. For traversal-path runs `planner_steps` is empty — we
    fall back to the step labels in `traces.steps[]`, which is the refined
    query for a pure traversal.
    """
    if not steps and traces:
        steps = [s.get("step", "") for s in (traces.get("steps") or []) if s.get("step")]
    if not steps:
        return ""
    return "\n".join(f"{i + 1}. {s}" for i, s in enumerate(steps))


# ── HTTP / SSE ───────────────────────────────────────────────────────────────

def _iter_sse_events(resp: requests.Response):
    """
    Yield (event_name, data_dict) for each SSE message on `resp`.

    Response must be opened with stream=True. Handles the canonical
    "event: <name>\\ndata: <json>\\n\\n" framing and skips `: heartbeat` lines.
    """
    current_event = ""
    current_data = ""
    for raw in resp.iter_lines(decode_unicode=True):
        if raw is None:
            continue
        line = raw.rstrip("\r")
        if line.startswith(":"):       # comment / heartbeat
            continue
        if line.startswith("event:"):
            current_event = line.split(":", 1)[1].strip()
        elif line.startswith("data:"):
            current_data = line.split(":", 1)[1].strip()
        elif line == "" and current_event:
            data: dict = {}
            if current_data:
                try:
                    data = json.loads(current_data)
                except json.JSONDecodeError:
                    data = {"raw": current_data}
            yield current_event, data
            current_event = ""
            current_data = ""


def call_simulate_stream(
    base_url: str,
    query: str,
    project_type: str,
    user_id: str,
    auto_resume_hitl: bool = True,
) -> dict:
    """
    Open GET /api/v1/simulate/stream, consume the SSE stream, and return the
    payload carried by the final `complete` event. Raises on error events or
    HTTP failure.

    On `hitl_start`, this fires POST /simulate/stream/resume with
    "Accept stated assumptions" and keeps reading the same stream.
    """
    thread_id = str(uuid.uuid4())
    params = {
        "query": query,
        "user_id": user_id,
        "project_type": project_type,
        "thread_id": thread_id,
    }
    url = base_url.rstrip("/") + "/api/v1/simulate/stream?" + urlencode(params)

    with requests.get(
        url,
        stream=True,
        timeout=(10, REQUEST_TIMEOUT_SECONDS),
        headers={"Accept": "text/event-stream"},
    ) as resp:
        resp.raise_for_status()

        for event_name, data in _iter_sse_events(resp):
            if event_name == "stream_started":
                thread_id = data.get("thread_id", thread_id)
            elif event_name == "hitl_start":
                if not auto_resume_hitl:
                    raise RuntimeError(
                        "Stream paused for clarification (hitl_start) and "
                        "--no-auto-resume-hitl was set"
                    )
                resume_url = base_url.rstrip("/") + "/api/v1/simulate/stream/resume"
                r = requests.post(
                    resume_url,
                    json={"thread_id": thread_id, "clarification": "Accept stated assumptions"},
                    timeout=30,
                )
                r.raise_for_status()
            elif event_name == "error":
                raise RuntimeError(f"stream error: {data.get('message', data)}")
            elif event_name == "complete":
                return data

    raise RuntimeError("stream ended without a 'complete' event")


# ── Workbook helpers ─────────────────────────────────────────────────────────

def load_or_create_workbook(path: Path) -> tuple[openpyxl.Workbook, "openpyxl.worksheet.worksheet.Worksheet", int]:
    """
    Open the workbook at `path` (create if missing).

    Returns (workbook, worksheet, next_id) where next_id is the ID to assign
    to the first new row.
    """
    if path.exists():
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        existing_headers = [c.value for c in ws[1]] if ws.max_row >= 1 else []
        if existing_headers[: len(HEADERS)] != HEADERS:
            raise SystemExit(
                f"Refusing to append — {path} exists but its header row is {existing_headers!r}, "
                f"not the expected {HEADERS!r}. Point --output at a fresh file or fix the header."
            )
        # Find max existing numeric ID in column A
        max_id = 0
        for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
            val = row[0]
            if isinstance(val, int):
                max_id = max(max_id, val)
            elif isinstance(val, str) and val.isdigit():
                max_id = max(max_id, int(val))
        return wb, ws, max_id + 1

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Simulation Results"
    ws.append(HEADERS)
    # Header styling
    header_fill = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for col_idx, name in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = COLUMN_WIDTHS[name]
    ws.freeze_panes = "A2"
    return wb, ws, 1


def append_row(ws, row_id: int, query: str, api_response: dict | None, error: str | None) -> None:
    """
    Write a result row. Binary/Feedback columns are left empty — reviewers
    fill those in manually after inspecting the model output.
    """
    if error is not None:
        planner_steps = ""
        response_text = f"ERROR: {error}"
        data_correct = f"ERROR: {error}"
        kpi_used = ""
        algorithm = ""
    else:
        api_response = api_response or {}
        traces = api_response.get("traces") or {}
        final_response = api_response.get("final_response", "") or ""
        planner_steps = format_planner_steps(api_response.get("planner_steps"), traces)
        response_text = final_response
        data_correct = final_response
        kpi_used = extract_kpi_used(traces)
        algorithm = extract_algorithm(api_response, traces)

    values_by_col: dict[int, Any] = {
        COL_ID: row_id,
        COL_QUERY: query,
        COL_PLANNER_STEPS: planner_steps,
        COL_RESPONSE: response_text,
        COL_DATA_CORRECT: data_correct,
        COL_KPI_USED: kpi_used,
        COL_ALGORITHM: algorithm,
    }

    new_row = ws.max_row + 1
    for col_idx in range(1, len(HEADERS) + 1):
        cell = ws.cell(row=new_row, column=col_idx, value=values_by_col.get(col_idx))
        cell.alignment = Alignment(wrap_text=True, vertical="top")


# ── Query file loading ───────────────────────────────────────────────────────

_VALID_PROJECT_TYPES = {"NTM", "AHLOB Modernization", "NTM,AHLOB Modernization"}


def load_queries(path: Path) -> list[tuple[str, str]]:
    """
    Parse a file where each non-blank, non-comment line is:
        "query text here"   "project_type"

    Uses shlex.split so quoted strings with spaces are one token each.
    Returns a list of (query, project_type) tuples, preserving order.
    """
    raw_lines = path.read_text(encoding="utf-8").splitlines()
    rows: list[tuple[str, str]] = []
    for lineno, line in enumerate(raw_lines, start=1):
        stripped = line.strip()
        if not stripped or stripped.startswith("#"):
            continue
        try:
            tokens = shlex.split(stripped)
        except ValueError as e:
            raise SystemExit(f"{path}:{lineno}: cannot parse line ({e}): {line!r}")
        if len(tokens) != 2:
            raise SystemExit(
                f"{path}:{lineno}: expected 2 quoted values (query, project_type), "
                f"got {len(tokens)}: {line!r}"
            )
        query, project_type = tokens[0].strip(), tokens[1].strip()
        if not query:
            raise SystemExit(f"{path}:{lineno}: empty query")
        if project_type not in _VALID_PROJECT_TYPES:
            raise SystemExit(
                f"{path}:{lineno}: project_type {project_type!r} is not one of "
                f"{sorted(_VALID_PROJECT_TYPES)}"
            )
        rows.append((query, project_type))
    if not rows:
        raise SystemExit(f"No queries found in {path}")
    return rows


# ── Main ─────────────────────────────────────────────────────────────────────

def run_batch(
    base_url: str,
    rows: list[tuple[str, str]],
    output: Path,
    concurrency: int,
    user_id: str,
    auto_resume_hitl: bool,
) -> None:
    wb, ws, next_id = load_or_create_workbook(output)
    print(f"→ writing to {output} (starting at ID {next_id})", flush=True)
    print(f"→ {len(rows)} queries, concurrency={concurrency}", flush=True)

    # Pre-assign IDs so rows stay in the same order as the input file,
    # regardless of which request finishes first.
    id_by_index = {i: next_id + i for i in range(len(rows))}
    results: dict[int, tuple[dict | None, str | None]] = {}

    write_lock = threading.Lock()
    completed = 0
    t0 = time.perf_counter()

    def worker(idx: int, query: str, project_type: str):
        try:
            body = call_simulate_stream(
                base_url, query, project_type, user_id, auto_resume_hitl=auto_resume_hitl
            )
            err = None
            if body.get("status") and body.get("status") != "complete":
                err = f"status={body.get('status')}"
            return idx, body, err
        except Exception as e:  # noqa: BLE001
            return idx, None, f"{type(e).__name__}: {e}"

    with ThreadPoolExecutor(max_workers=concurrency) as pool:
        futures = [pool.submit(worker, i, q, pt) for i, (q, pt) in enumerate(rows)]
        for fut in as_completed(futures):
            idx, body, err = fut.result()
            results[idx] = (body, err)
            with write_lock:
                completed += 1
                tag = "✓" if err is None else "✗"
                q, pt = rows[idx]
                preview = q[:55].replace("\n", " ")
                print(
                    f"  [{completed}/{len(rows)}] {tag} id={id_by_index[idx]}  "
                    f"[{pt}]  {preview}{' …' if len(q) > 55 else ''}"
                    + (f"   ({err})" if err else ""),
                    flush=True,
                )

    # Write rows in original input order so IDs stay consecutive
    for idx in range(len(rows)):
        body, err = results.get(idx, (None, "no result"))
        append_row(ws, id_by_index[idx], rows[idx][0], body, err)

    wb.save(output)

    elapsed = time.perf_counter() - t0
    ok = sum(1 for v in results.values() if v[1] is None)
    print(f"\n✓ done in {elapsed:.1f}s — {ok}/{len(rows)} succeeded → {output}", flush=True)


def parse_args(argv: list[str]) -> argparse.Namespace:
    p = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    p.add_argument("--base-url", default="http://localhost:8000", help="API base URL (default: %(default)s)")
    p.add_argument("--queries", required=True, type=Path,
                   help='Path to text file. Each line: "query" "project_type"')
    p.add_argument("--output", required=True, type=Path, help="Path to the target .xlsx file")
    p.add_argument("--concurrency", type=int, default=5, help="Max in-flight SSE streams (default: 5)")
    p.add_argument("--user-id", default="batch-runner", help="user_id passed to the API")
    p.add_argument("--no-auto-resume-hitl", action="store_true",
                   help="Do NOT auto-resume on hitl_start; mark the row as errored instead")
    return p.parse_args(argv)


def main(argv: list[str] | None = None) -> int:
    args = parse_args(argv or sys.argv[1:])
    if args.concurrency < 1:
        raise SystemExit("--concurrency must be >= 1")
    rows = load_queries(args.queries)
    run_batch(
        base_url=args.base_url,
        rows=rows,
        output=args.output,
        concurrency=args.concurrency,
        user_id=args.user_id,
        auto_resume_hitl=not args.no_auto_resume_hitl,
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
