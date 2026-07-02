"""
Run every question in queries.txt FOUR times — each run on its own thread_id —
against the SSE streaming endpoint, and record the four runs side-by-side in a
wide .xlsx sheet for manual consistency/accuracy review.

Why 4 runs on 4 threads
------------------------
The same question is answered four independent times so a reviewer can eyeball
how consistent the agent is (does it give the same numbers / logic / queries
each time?). Each run gets a fresh thread_id so there is no checkpointer bleed
between runs.

Sheet layout (row 1 = headers, row 2 = descriptor legend, data from row 3)
--------------------------------------------------------------------------
    Question |
    [ Response 1 | Logic 1 | Data 1 | Traversal 1 |
      Response validation | Logic validation | Data validation |
      Traversal validation | Thread 1 ]        <-- repeated for runs 2, 3, 4
    | Consistency | Accuracy

Per-run data cells the script fills (validation + Consistency + Accuracy are
left blank for manual review):
    Response   = final_response
    Logic      = execution_algorithm  +  planner_steps
    Data       = the Cypher / SQL-python queries executed (from traces)
    Traversal  = the full tool_calls list (name, input, output, status)
    Thread     = the thread_id used for that run

HITL handling
-------------
If a run pauses for clarification (hitl_start), the script auto-resumes that
run's own thread with the answer "across all regions" and keeps reading the
same stream.

Ordering
--------
Questions are processed one at a time. All four runs of a question finish
(barrier) before the row is written, the workbook is saved, and the next
question starts — so the sheet fills in progressively and safely.

Endpoints used
--------------
    GET  /api/v1/simulate/stream          (query, user_id, project_type, thread_id)
    POST /api/v1/simulate/stream/resume   (thread_id, clarification)

Usage
-----
    venv/bin/python scripts/run_consistency_sheet.py \
        --base-url http://localhost:8000 \
        --queries  queries.txt \
        --output   consistency_results.xlsx
"""
from __future__ import annotations

import argparse
import json
import shlex
import sys
import time
import uuid
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path
from typing import Any
from urllib.parse import urlencode

import openpyxl
import requests
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


# ── Configuration ────────────────────────────────────────────────────────────

NUM_RUNS = 4                                # runs per question (== sheet columns)
HITL_CLARIFICATION = "across all regions"   # answer sent on any hitl_start
REQUEST_TIMEOUT_SECONDS = 900               # a single simulation can run minutes
CELL_LIMIT = 32000                          # Excel hard-caps a cell at 32,767 chars

# Endpoint base on the running server (port 8000). The stream endpoint is
# <API_BASE>/stream and resume is <API_BASE>/stream/resume. The live "Simulation
# Agent API" mounts these under /simulate. Override with --api-base if a server
# mounts the router elsewhere (e.g. an /analyze deployment).
DEFAULT_API_BASE = "/api/v1/simulate"

_VALID_PROJECT_TYPES = {"NTM", "AHLOB Modernization", "NTM,AHLOB Modernization", "Both", "NAS"}

# Per-run block: 4 data fields + 4 manual validation fields + thread id = 9 cols.
# Column offsets are 0-based within a run block.
OFF_RESPONSE   = 0
OFF_LOGIC      = 1
OFF_DATA       = 2
OFF_TRAVERSAL  = 3
# offsets 4..7 → the four *validation* columns (left blank)
OFF_THREAD     = 8
RUN_BLOCK_WIDTH = 9

COL_QUESTION = 1                                   # column A
FIRST_RUN_COL = 2                                  # run 1 starts at column B
COL_CONSISTENCY = FIRST_RUN_COL + NUM_RUNS * RUN_BLOCK_WIDTH       # 38
COL_ACCURACY    = COL_CONSISTENCY + 1                              # 39
TOTAL_COLS      = COL_ACCURACY                                     # 39


def _run_base_col(run_idx: int) -> int:
    """1-based column where run `run_idx` (0-based) block starts."""
    return FIRST_RUN_COL + run_idx * RUN_BLOCK_WIDTH


# ── Header construction ──────────────────────────────────────────────────────

def build_headers() -> list[str]:
    headers = ["Question"]
    for i in range(1, NUM_RUNS + 1):
        headers += [
            f"Response {i}", f"Logic {i}", f"Data {i}", f"Traversal {i}",
            "Response validation", "Logic validation",
            "Data validation", "Traversal validation",
            f"Thread {i}",
        ]
    headers += ["Consistency", "Accuracy"]
    return headers


def build_descriptor() -> list[str]:
    """Row 2 — the legend that documents what each cell holds."""
    desc = [""]
    for _ in range(NUM_RUNS):
        desc += [
            "final_response", "algorithm + planner steps", "Queries", "tool_calls",
            "bool", "bool", "bool", "bool",
            "",                     # Thread column has no type legend
        ]
    desc += ["", ""]               # Consistency, Accuracy
    return desc


HEADERS = build_headers()
DESCRIPTOR = build_descriptor()

COLUMN_WIDTHS = {"Question": 45, "Consistency": 14, "Accuracy": 14}
for _i in range(1, NUM_RUNS + 1):
    COLUMN_WIDTHS.update({
        f"Response {_i}": 55, f"Logic {_i}": 50, f"Data {_i}": 55,
        f"Traversal {_i}": 60, f"Thread {_i}": 38,
    })
# validation headers all share the same width; set below when building the sheet
_VALIDATION_WIDTH = 12


# ── Cell helpers ─────────────────────────────────────────────────────────────

def _cap(text: str) -> str:
    """Keep a cell under Excel's 32,767 char limit."""
    if text is None:
        return ""
    text = str(text)
    if len(text) <= CELL_LIMIT:
        return text
    return text[:CELL_LIMIT] + "\n…[truncated]"


def _iter_tool_calls(traces: dict | None):
    """Yield (step_label, tool_call) across every planner/traversal step, in order."""
    if not traces:
        return
    for step in traces.get("steps", []) or []:
        label = step.get("step", "")
        for tc in step.get("tool_calls", []) or []:
            yield label, tc


# ── Field extraction from the `complete` SSE payload ──────────────────────────

def extract_response(data: dict) -> str:
    return _cap(data.get("final_response", "") or "")


def extract_logic(data: dict) -> str:
    """Logic column = execution_algorithm narrative + the planner steps."""
    algo = (data.get("execution_algorithm") or "").strip()
    steps = data.get("planner_steps") or []
    traces = data.get("traces") or {}
    if not steps:  # traversal path has no planner steps → use trace step labels
        steps = [s.get("step", "") for s in (traces.get("steps") or []) if s.get("step")]

    parts: list[str] = []
    if algo:
        parts.append("ALGORITHM:\n" + algo)
    if steps:
        parts.append(
            "PLANNER STEPS:\n"
            + "\n".join(f"{i + 1}. {s}" for i, s in enumerate(steps))
        )
    return _cap("\n\n".join(parts))


def extract_data_queries(data: dict) -> str:
    """
    Data column = the queries the agent actually ran to fetch data:
    run_cypher (Cypher) and run_sql_python / run_python (code that queries the KG).
    """
    traces = data.get("traces") or {}
    blocks: list[str] = []
    n = 0
    for step_label, tc in _iter_tool_calls(traces):
        name = tc.get("tool_name", "")
        ti = tc.get("tool_input") or {}
        snippet = ""
        if name == "run_cypher":
            snippet = ti.get("query", "")
        elif name in ("run_sql_python", "run_python"):
            snippet = ti.get("code", "")
        if not snippet:
            continue
        n += 1
        header = f"── [{n}] {name}" + (f"  (step: {step_label})" if step_label else "")
        blocks.append(f"{header}\n{snippet}")
    return _cap("\n\n".join(blocks))


def extract_traversal(data: dict) -> str:
    """Traversal column = the full tool_calls trace (name, input, output, status)."""
    traces = data.get("traces") or {}
    lines: list[str] = []
    for step in traces.get("steps", []) or []:
        lines.append(f"━━━━━ STEP: {step.get('step', '')}")
        for j, tc in enumerate(step.get("tool_calls", []) or [], start=1):
            name = tc.get("tool_name", "")
            status = tc.get("status", "")
            ms = tc.get("execution_time_ms", "")
            inp = json.dumps(tc.get("tool_input", {}), default=str, ensure_ascii=False)
            out = tc.get("tool_output")
            out_str = out if isinstance(out, str) else json.dumps(out, default=str, ensure_ascii=False)
            lines.append(f"[{j}] {name}  ({status}, {ms}ms)")
            lines.append(f"    input:  {inp}")
            lines.append(f"    output: {out_str}")
        lines.append("")
    return _cap("\n".join(lines))


# ── SSE plumbing ─────────────────────────────────────────────────────────────

def _iter_sse_events(resp: requests.Response):
    """Yield (event_name, data_dict) for each SSE message. resp must be stream=True."""
    current_event = ""
    current_data = ""
    for raw in resp.iter_lines(decode_unicode=True):
        if raw is None:
            continue
        line = raw.rstrip("\r")
        if line.startswith(":"):          # heartbeat / comment
            continue
        if line.startswith("event:"):
            current_event = line.split(":", 1)[1].strip()
        elif line.startswith("data:"):
            current_data = line.split(":", 1)[1].strip()
        elif line == "" and current_event:
            data: dict = {}
            if current_data:
                try:
                    data = json.loads(current_data)
                except json.JSONDecodeError:
                    data = {"raw": current_data}
            yield current_event, data
            current_event = ""
            current_data = ""


def run_single_stream(
    base_url: str,
    api_base: str,
    query: str,
    project_type: str,
    user_id: str,
) -> dict:
    """
    Open one SSE stream, auto-resume any HITL pause with HITL_CLARIFICATION on
    this run's own thread, and return:
        {"thread_id": str, "data": dict|None, "error": str|None}
    `data` is the payload of the final `complete` event.
    """
    thread_id = str(uuid.uuid4())
    params = {
        "query": query,
        "user_id": user_id,
        "project_type": project_type,
        "thread_id": thread_id,
    }
    stream_url = base_url.rstrip("/") + "/" + api_base.strip("/") + "/stream"
    resume_url = stream_url + "/resume"
    url = stream_url + "?" + urlencode(params)

    try:
        with requests.get(
            url,
            stream=True,
            timeout=(10, REQUEST_TIMEOUT_SECONDS),
            headers={"Accept": "text/event-stream"},
        ) as resp:
            resp.raise_for_status()
            for event_name, edata in _iter_sse_events(resp):
                if event_name == "stream_started":
                    thread_id = edata.get("thread_id", thread_id)
                elif event_name == "hitl_start":
                    r = requests.post(
                        resume_url,
                        json={"thread_id": thread_id, "clarification": HITL_CLARIFICATION},
                        timeout=30,
                    )
                    r.raise_for_status()
                elif event_name == "error":
                    return {"thread_id": thread_id, "data": None,
                            "error": f"stream error: {edata.get('message', edata)}"}
                elif event_name == "complete":
                    return {"thread_id": thread_id, "data": edata, "error": None}
        return {"thread_id": thread_id, "data": None,
                "error": "stream ended without a 'complete' event"}
    except Exception as e:  # noqa: BLE001
        return {"thread_id": thread_id, "data": None, "error": f"{type(e).__name__}: {e}"}


def run_question(
    base_url: str,
    api_base: str,
    query: str,
    project_type: str,
    user_id: str,
    run_concurrency: int,
) -> list[dict]:
    """Run the question NUM_RUNS times; return results in run order (0..NUM_RUNS-1)."""
    results: list[dict | None] = [None] * NUM_RUNS
    with ThreadPoolExecutor(max_workers=run_concurrency) as pool:
        fut_to_idx = {
            pool.submit(run_single_stream, base_url, api_base, query, f"{project_type}", user_id): i
            for i in range(NUM_RUNS)
        }
        for fut in as_completed(fut_to_idx):
            idx = fut_to_idx[fut]
            results[idx] = fut.result()
    return [r or {"thread_id": "", "data": None, "error": "no result"} for r in results]


# ── Workbook ─────────────────────────────────────────────────────────────────

def load_or_create_workbook(path: Path):
    """Return (wb, ws, existing_questions:set[str]). Creates the sheet if missing."""
    if path.exists():
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        existing_headers = [c.value for c in ws[1]] if ws.max_row >= 1 else []
        if existing_headers[: len(HEADERS)] != HEADERS:
            raise SystemExit(
                f"Refusing to append — {path} exists but its header row does not "
                f"match this script's layout. Point --output at a fresh file."
            )
        existing = set()
        # data rows start at row 3 (row 1 header, row 2 descriptor)
        for row in ws.iter_rows(min_row=3, max_col=1, values_only=True):
            if row[0]:
                existing.add(str(row[0]).strip())
        return wb, ws, existing

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Consistency Results"

    header_fill = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
    desc_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    desc_font = Font(italic=True, color="333333")

    ws.append(HEADERS)
    ws.append(DESCRIPTOR)
    for col_idx, name in enumerate(HEADERS, start=1):
        letter = get_column_letter(col_idx)
        hc = ws.cell(row=1, column=col_idx)
        hc.fill = header_fill
        hc.font = header_font
        hc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        dc = ws.cell(row=2, column=col_idx)
        dc.fill = desc_fill
        dc.font = desc_font
        dc.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[letter].width = COLUMN_WIDTHS.get(name, _VALIDATION_WIDTH)
    ws.freeze_panes = "B3"     # freeze the Question column + both header rows
    return wb, ws, set()


def write_question_row(ws, query: str, run_results: list[dict]) -> None:
    """Append one row: the question + 4 run blocks. Validation cells left blank."""
    new_row = ws.max_row + 1
    values_by_col: dict[int, Any] = {COL_QUESTION: query}

    for run_idx, res in enumerate(run_results):
        base = _run_base_col(run_idx)
        data = res.get("data")
        err = res.get("error")
        thread_id = res.get("thread_id", "")

        if data is None:
            marker = f"ERROR: {err}"
            values_by_col[base + OFF_RESPONSE] = marker
            values_by_col[base + OFF_LOGIC] = ""
            values_by_col[base + OFF_DATA] = ""
            values_by_col[base + OFF_TRAVERSAL] = ""
        else:
            values_by_col[base + OFF_RESPONSE] = extract_response(data)
            values_by_col[base + OFF_LOGIC] = extract_logic(data)
            values_by_col[base + OFF_DATA] = extract_data_queries(data)
            values_by_col[base + OFF_TRAVERSAL] = extract_traversal(data)
        values_by_col[base + OFF_THREAD] = thread_id
        # validation offsets 4..7 intentionally left blank (manual review)

    for col_idx in range(1, TOTAL_COLS + 1):
        cell = ws.cell(row=new_row, column=col_idx, value=values_by_col.get(col_idx))
        cell.alignment = Alignment(wrap_text=True, vertical="top")


# ── Query file parsing ───────────────────────────────────────────────────────

def load_queries(path: Path) -> list[tuple[str, str]]:
    """
    Each non-blank, non-comment line: "query text"  "project_type"
    An optional leading integer index token (e.g. `1  "q"  "NTM"`) is tolerated.
    """
    raw_lines = path.read_text(encoding="utf-8").splitlines()
    rows: list[tuple[str, str]] = []
    for lineno, line in enumerate(raw_lines, start=1):
        stripped = line.strip()
        if not stripped or stripped.startswith("#"):
            continue
        try:
            tokens = shlex.split(stripped)
        except ValueError as e:
            raise SystemExit(f"{path}:{lineno}: cannot parse line ({e}): {line!r}")
        # tolerate a leading line-number token
        if len(tokens) == 3 and tokens[0].isdigit():
            tokens = tokens[1:]
        if len(tokens) != 2:
            raise SystemExit(
                f"{path}:{lineno}: expected 2 quoted values (query, project_type), "
                f"got {len(tokens)}: {line!r}"
            )
        query, project_type = tokens[0].strip(), tokens[1].strip()
        if not query:
            raise SystemExit(f"{path}:{lineno}: empty query")
        if project_type not in _VALID_PROJECT_TYPES:
            raise SystemExit(
                f"{path}:{lineno}: project_type {project_type!r} is not one of "
                f"{sorted(_VALID_PROJECT_TYPES)}"
            )
        rows.append((query, project_type))
    if not rows:
        raise SystemExit(f"No queries found in {path}")
    return rows


# ── Orchestration ────────────────────────────────────────────────────────────

def run_batch(
    base_url: str,
    api_base: str,
    rows: list[tuple[str, str]],
    output: Path,
    run_concurrency: int,
    user_id: str,
    skip_existing: bool,
) -> None:
    wb, ws, existing = load_or_create_workbook(output)
    print(f"→ output: {output}", flush=True)
    print(f"→ endpoint: {base_url.rstrip('/')}/{api_base.strip('/')}/stream", flush=True)
    print(f"→ {len(rows)} question(s), {NUM_RUNS} runs each, "
          f"run_concurrency={run_concurrency}", flush=True)
    if existing:
        print(f"→ sheet already has {len(existing)} question(s)", flush=True)

    t_all = time.perf_counter()
    for q_idx, (query, project_type) in enumerate(rows, start=1):
        if skip_existing and query.strip() in existing:
            print(f"\n[{q_idx}/{len(rows)}] SKIP (already in sheet): {query[:70]}", flush=True)
            continue

        print(f"\n[{q_idx}/{len(rows)}] [{project_type}] {query[:70]}"
              f"{' …' if len(query) > 70 else ''}", flush=True)
        t0 = time.perf_counter()

        results = run_question(base_url, api_base, query, project_type, user_id, run_concurrency)

        # ── All 4 runs are done here (barrier). Report, then persist the row. ──
        for i, r in enumerate(results, start=1):
            if r.get("error"):
                print(f"    run {i}: ✗ {r['error']}  (thread={r.get('thread_id','')[:8]})", flush=True)
            else:
                d = r["data"]
                routing = d.get("routing_decision", "?")
                n_tools = (d.get("traces") or {}).get("total_tool_calls", 0)
                print(f"    run {i}: ✓ {routing}, {n_tools} tool calls  "
                      f"(thread={r.get('thread_id','')[:8]})", flush=True)

        write_question_row(ws, query, results)
        wb.save(output)                        # persist after every question
        existing.add(query.strip())
        print(f"    → row saved ({time.perf_counter() - t0:.1f}s)", flush=True)

    print(f"\n✓ done in {time.perf_counter() - t_all:.1f}s → {output}", flush=True)


def parse_args(argv: list[str]) -> argparse.Namespace:
    p = argparse.ArgumentParser(
        description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter
    )
    p.add_argument("--base-url", default="http://localhost:8000", help="API base URL")
    p.add_argument("--api-base", default=DEFAULT_API_BASE,
                   help=f"Router base path; stream=<api-base>/stream (default: {DEFAULT_API_BASE})")
    p.add_argument("--queries", type=Path, default=Path("queries.txt"),
                   help='Query file. Each line: "query" "project_type"')
    p.add_argument("--output", type=Path, default=Path("consistency_results.xlsx"),
                   help="Target .xlsx file")
    p.add_argument("--run-concurrency", type=int, default=NUM_RUNS,
                   help=f"How many of the {NUM_RUNS} runs to fire at once (default: {NUM_RUNS})")
    p.add_argument("--user-id", default="consistency-runner", help="user_id passed to the API")
    p.add_argument("--no-skip-existing", dest="skip_existing", action="store_false",
                   help="Re-run questions even if they are already in the sheet")
    p.set_defaults(skip_existing=True)
    return p.parse_args(argv)


def main(argv: list[str] | None = None) -> int:
    args = parse_args(argv or sys.argv[1:])
    if not (1 <= args.run_concurrency <= NUM_RUNS):
        raise SystemExit(f"--run-concurrency must be between 1 and {NUM_RUNS}")
    rows = load_queries(args.queries)
    run_batch(
        base_url=args.base_url,
        api_base=args.api_base,
        rows=rows,
        output=args.output,
        run_concurrency=args.run_concurrency,
        user_id=args.user_id,
        skip_existing=args.skip_existing,
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
