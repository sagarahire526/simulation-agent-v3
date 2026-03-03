"""
One-time ingestion script: reads Nokia Scenario Excel and inserts rows with
embeddings into pwc_semantic_information_schema.semantics_simulation.

Embedding is created by concatenating the Scenario text and Data Phase Questions,
then calling OpenAI text-embedding-3-small.

Usage (from the simulation-agent-v1 directory):
    ../venv/bin/python scripts/ingest_scenarios.py
    # or
    venv/bin/python scripts/ingest_scenarios.py
"""
from __future__ import annotations

import os
import re
import sys

# Ensure the project root is importable
_PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, _PROJECT_ROOT)

from dotenv import load_dotenv
load_dotenv(os.path.join(_PROJECT_ROOT, ".env"))

import openpyxl
import psycopg2

import config

EXCEL_PATH = os.path.join(_PROJECT_ROOT, "Copy of Nokia Scenario.xlsx")
SHEET_NAME = "Simulator Samples"
TABLE = "pwc_semantic_information_schema.semantics_simulation"
EMBEDDING_MODEL = "text-embedding-3-small"


# ── Helpers ────────────────────────────────────────────────────────────────

def _parse_text_to_array(text: str | None) -> list[str]:
    """Split multi-paragraph text (separated by blank lines) into list items."""
    if not text:
        return []
    items = [item.strip() for item in re.split(r"\n\s*\n", text) if item.strip()]
    return items


def _create_embedding(client, text: str) -> list[float]:
    """Call OpenAI text-embedding-3-small to get a float vector."""
    response = client.embeddings.create(
        model=EMBEDDING_MODEL,
        input=text.strip(),
    )
    return response.data[0].embedding


def _build_embedding_text(scenario: str | None, data_phase_questions: list[str]) -> str:
    """Concatenate Scenario and Data Phase Questions for embedding."""
    parts = [scenario or ""]
    if data_phase_questions:
        parts.append("\n".join(data_phase_questions))
    return "\n\n".join(filter(None, parts))


# ── Main ingestion ─────────────────────────────────────────────────────────

def ingest(clear_existing: bool = False) -> None:
    """
    Read the Excel file and insert/update rows in semantics_simulation.

    Args:
        clear_existing: If True, truncates the table before inserting.
                        Useful when re-running with updated data.
    """
    from openai import OpenAI
    openai_client = OpenAI(api_key=config.OPENAI_API_KEY)

    print(f"📂 Loading Excel: {EXCEL_PATH}")
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb[SHEET_NAME]
    print(f"   Sheet '{SHEET_NAME}': {ws.max_row - 1} potential data rows\n")

    conn = psycopg2.connect(
        host=config.PG_HOST,
        port=config.PG_PORT,
        database=config.PG_DATABASE,
        user=config.PG_USER,
        password=config.PG_PASSWORD,
    )
    cur = conn.cursor()

    if clear_existing:
        print("🗑  Clearing existing rows from semantics_simulation...")
        cur.execute(f"TRUNCATE TABLE {TABLE}")
        conn.commit()

    inserted = updated = skipped = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        (
            scenario_id_raw,
            scenario,
            data_phase_steps_txt,
            data_phase_questions_txt,
            calc_steps_txt,
            sim_steps_txt,
            methodology,
        ) = row[:7]

        # Skip empty rows
        if scenario_id_raw is None:
            continue

        scenario_id = int(scenario_id_raw)

        # Parse array fields from multi-paragraph text
        data_phase_steps = _parse_text_to_array(data_phase_steps_txt)
        data_phase_questions = _parse_text_to_array(data_phase_questions_txt)
        calc_steps = _parse_text_to_array(calc_steps_txt)
        sim_steps = _parse_text_to_array(sim_steps_txt)

        # Build embedding text: Scenario + Data Phase Questions
        emb_text = _build_embedding_text(scenario, data_phase_questions)

        print(f"  🔢 Creating embedding for Scenario {scenario_id}...", end=" ", flush=True)
        embedding = _create_embedding(openai_client, emb_text)
        print(f"✓  ({len(embedding)}-dim vector)")

        # Upsert — insert if not exists, update if exists
        cur.execute(f"SELECT 1 FROM {TABLE} WHERE scenario_id = %s", (scenario_id,))
        exists = cur.fetchone()

        if exists:
            cur.execute(
                f"""
                UPDATE {TABLE} SET
                    scenario = %s,
                    data_phase_steps = %s,
                    data_phase_questions = %s,
                    calculation_phase_steps = %s,
                    simulator_phase_steps = %s,
                    simulation_methodology = %s,
                    embedding = %s,
                    updated_at = NOW()
                WHERE scenario_id = %s
                """,
                (
                    scenario, data_phase_steps, data_phase_questions,
                    calc_steps, sim_steps, methodology, embedding, scenario_id,
                ),
            )
            updated += 1
            print(f"     ↻  Scenario {scenario_id} updated.")
        else:
            cur.execute(
                f"""
                INSERT INTO {TABLE}
                    (scenario_id, scenario, data_phase_steps, data_phase_questions,
                     calculation_phase_steps, simulator_phase_steps,
                     simulation_methodology, embedding, created_at, updated_at)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, NOW(), NOW())
                """,
                (
                    scenario_id, scenario, data_phase_steps, data_phase_questions,
                    calc_steps, sim_steps, methodology, embedding,
                ),
            )
            inserted += 1
            print(f"     ✅ Scenario {scenario_id} inserted.")

        conn.commit()

    cur.close()
    conn.close()

    print(f"\n{'─' * 50}")
    print(f"✅ Ingestion complete: {inserted} inserted, {updated} updated, {skipped} skipped.")


if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description="Ingest Nokia Scenario Excel into semantics_simulation table.")
    parser.add_argument(
        "--clear",
        action="store_true",
        help="Truncate the table before ingesting (fresh reload).",
    )
    args = parser.parse_args()
    ingest(clear_existing=args.clear)
