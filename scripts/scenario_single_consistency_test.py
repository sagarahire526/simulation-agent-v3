"""
scripts/scenario_single_consistency_test.py

DEEP single-scenario consistency test. Unlike scenario_consistency_test.py (which sweeps
all 9 scenarios shallowly), this drills into ONE scenario and runs each case 4x to prove
two things a program office actually cares about:

  BLOCK A — SAME scenario, DIFFERENT params (execution consistency)
      Several param sets (default / region SOUTH / region WEST / alt breakdown / alt window),
      each executed 4x. Asserts: byte-identical across the 4 runs of a set (determinism),
      and different param sets produce different output (params actually bite).

  BLOCK B — SAME scenario, DIFFERENT question framings (full production path)
      Paraphrases of the same underlying ask, each run 4x through the ENTIRE production
      path: search_scenarios() embedding match  ->  LLM param extraction  ->  run_scenario().
      Asserts: every framing matches the SAME scenario node, matching is stable across the
      4 runs, extracted params are stable, and execution is byte-identical across the 4 runs.

Block A hits Postgres via the isolated node runner (no matching/LLM). Block B is the real
end-to-end deterministic path minus the graph plumbing, so it also catches "a rephrase
routes to the wrong scenario" and "the extractor is jittery on this phrasing".

Everything is CURRENT_DATE-relative, so "identical" means same-day, same-input.

Usage:
  venv/bin/python scripts/scenario_single_consistency_test.py                        # default scn-001
  venv/bin/python scripts/scenario_single_consistency_test.py --scenario scn-007-po-cohort-plan
  venv/bin/python scripts/scenario_single_consistency_test.py --runs 4 --out scn001_deep.xlsx
"""
from __future__ import annotations

import argparse
import hashlib
import json
import os
import sys
from datetime import date

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import services.sandbox_service as sandbox_svc  # noqa: E402
from tools.neo4j_tool import Neo4jTool  # noqa: E402

TEST_SMP = "NTM"
TEST_PROJECT_TYPE = "NTM"
DEFAULT_SCENARIO = "scn-001-scop-acceptance-prediction"

# Genuine paraphrases per scenario — used by Block B to test that a rephrase still routes
# to the SAME node and extracts stable params. Falls back to [canonical question] x xN if a
# scenario isn't listed here (still tests match+extract determinism on the canonical text).
FRAMINGS: dict[str, list[str]] = {
    "scn-001-scop-acceptance-prediction": [
        "Predict the SCOP acceptance date for each pending site",
        "When are the pending sites expected to get SCOP acceptance?",
        "Forecast the SCOP sign-off timing for sites still awaiting acceptance",
        "Give me the expected SCOP acceptance dates for the pending sites",
    ],
    "scn-007-po-cohort-plan": [
        "500 sites were allotted but a PO was raised for only 300 — plan the buildable cohort",
        "We have POs for just 300 of the 500 allocated sites, what does the plan look like?",
        "Split the 500 allocated sites by whether a PO exists and plan accordingly",
        "Only 300 of 500 sites have a purchase order — give me the cohort plan",
    ],
    "scn-008-crew-shortage-plan": [
        "If we have a 15% crew shortage in the South, what is the revised weekly Cx-start run rate?",
        "A 15 percent crew shortfall in South region — how does it hit our weekly construction-start rate?",
        "With 15% fewer crews in the South, what's the impact on our weekly run rate and sites at risk?",
        "Model a 15% crew reduction in South and show the revised weekly start capacity",
    ],
}


def _canon(obj) -> str:
    return json.dumps(obj, sort_keys=True, default=str)


def _hash(obj) -> str:
    return hashlib.sha256(_canon(obj).encode()).hexdigest()[:12]


def _fetch_scenario(scn_id: str) -> dict:
    t = Neo4jTool()
    try:
        r = t.run_cypher(
            "MATCH (n:BKGNode {node_id:$nid, entity_type:'scenario'}) "
            "RETURN n.node_id AS node_id, n.label AS label, "
            "coalesce(n.scn_canonical_question,'') AS q, "
            "coalesce(n.scn_param_schema,'') AS schema",
            {"nid": scn_id},
        )
    finally:
        t.close()
    recs = r.get("records", []) if r.get("status") == "success" else []
    if not recs:
        raise SystemExit(f"Scenario node '{scn_id}' not found.")
    rec = recs[0]
    schema = None
    try:
        schema = json.loads(rec["schema"]) if rec["schema"].strip() else None
    except json.JSONDecodeError:
        schema = None
    return {"node_id": rec["node_id"], "label": rec["label"],
            "question": rec["q"], "schema": schema}


# --------------------------------------------------------------------------- Block A


def _legacy_window(months_back: int) -> tuple[str, str]:
    today = date.today()
    y, m = today.year, today.month - months_back
    while m <= 0:
        m += 12
        y -= 1
    return date(y, m, 1).isoformat(), today.isoformat()


def _param_variants(scn: dict) -> list[dict]:
    """Build labelled (filter, group_by) param sets for Block A from the scenario schema.
    Legacy scn-001 (no 'fields') is handled with its start_date/end_date window."""
    schema = scn["schema"] or {}
    fields = schema.get("fields")
    if not fields:
        s2, e = _legacy_window(2)
        s4, _ = _legacy_window(4)
        base = {"smp_name": TEST_SMP}
        return [
            {"label": "default (2mo, no region)", "filter": {**base, "start_date": s2, "end_date": e}, "group_by": "construction_gc"},
            {"label": "region=SOUTH", "filter": {**base, "start_date": s2, "end_date": e, "rgn_region": "SOUTH"}, "group_by": "construction_gc"},
            {"label": "region=WEST", "filter": {**base, "start_date": s2, "end_date": e, "rgn_region": "WEST"}, "group_by": "construction_gc"},
            {"label": "group_by=m_market", "filter": {**base, "start_date": s2, "end_date": e}, "group_by": "m_market"},
            {"label": "window=4mo", "filter": {**base, "start_date": s4, "end_date": e}, "group_by": "construction_gc"},
        ]

    # Schema-driven: start from defaults, then vary region / breakdown.
    gbf = schema.get("group_by_field")
    base_filter, base_gb = {}, None
    for f in fields:
        nm = f.get("name")
        if not nm:
            continue
        dv = f.get("default")
        if nm == gbf:
            base_gb = dv
        elif dv not in (None, "", []):
            base_filter[nm] = dv
    base_filter["smp_name"] = TEST_SMP
    base_filter["project_type"] = TEST_PROJECT_TYPE

    variants = [{"label": "default scope", "filter": dict(base_filter), "group_by": base_gb}]
    variants.append({"label": "region=SOUTH", "filter": {**base_filter, "rgn_region": "SOUTH"}, "group_by": base_gb})
    variants.append({"label": "region=WEST", "filter": {**base_filter, "rgn_region": "WEST"}, "group_by": base_gb})
    # alternate breakdown, if the group_by field enumerates more than one option
    gb_field = next((f for f in fields if f.get("name") == gbf), None)
    if gb_field:
        allowed = gb_field.get("allowed") or []
        alt = next((a for a in allowed if a != base_gb), None)
        if alt is not None:
            variants.append({"label": f"group_by={alt}", "filter": dict(base_filter), "group_by": alt})
    return variants


def _run(scn_id, filt, group_by, timeout):
    out = sandbox_svc.run_bkg_node("scenario", scn_id, filters=filt, group_by=group_by, timeout_seconds=timeout)
    if out.get("status") != "success":
        return None, str(out.get("error"))[:300]
    return out.get("result"), None


def _metric(result) -> str:
    if not isinstance(result, dict):
        return str(result)[:80]
    keys = ("target_sites", "sites_at_risk", "revised_weekly_run_rate", "pending_count",
            "total_crews", "po_gate", "cohorts", "predictions", "horizon_weeks")
    bits = []
    for k in keys:
        if k in result:
            v = result[k]
            if isinstance(v, (list, dict)):
                v = f"<{len(v)}>"
            bits.append(f"{k}={v}")
    return ", ".join(bits)[:140] or ", ".join(list(result.keys())[:6])


def block_a(scn, runs, timeout):
    print(f"== BLOCK A: {scn['node_id']} — different params, {runs} runs each ==")
    rows, hashes = [], {}
    for v in _param_variants(scn):
        canon, identical, err, first = None, True, None, None
        for _ in range(runs):
            res, e = _run(scn["node_id"], v["filter"], v["group_by"], timeout)
            if e:
                err, identical = e, False
                break
            if first is None:
                first = res
            c = _canon(res)
            if canon is None:
                canon = c
            elif c != canon:
                identical = False
        hashes[v["label"]] = _hash(canon) if canon else None
        rows.append({
            "param_set": v["label"], "group_by": v["group_by"],
            "filter": _canon(v["filter"]), "runs": runs,
            "deterministic": "ERROR" if err else ("PASS" if identical else "FAIL"),
            "result_hash": hashes[v["label"]] or "",
            "metric": _metric(first) if first is not None else "",
            "error": err or "",
        })
        print(f"  [{rows[-1]['deterministic']:4}] {v['label']:26} hash={hashes[v['label']]}  {rows[-1]['metric'][:60]}")
    distinct = len({h for h in hashes.values() if h})
    print(f"  -> {distinct} distinct result(s) across {len(hashes)} param sets "
          f"({'params bite' if distinct > 1 else 'WARNING: all identical'})\n")
    return rows, distinct, len(hashes)


# --------------------------------------------------------------------------- Block B


def _match(query):
    from services.schema_embedding_service import search_scenarios
    return search_scenarios(query, project_type=TEST_PROJECT_TYPE)


def _extract(query, scn):
    from services.scenario_params import extract_params_by_schema, extract_scenario_params
    if scn["schema"] and scn["schema"].get("fields"):
        return extract_params_by_schema(query, scn["schema"], project_type=TEST_PROJECT_TYPE)
    return extract_scenario_params(query, project_type=TEST_PROJECT_TYPE)


def block_b(scn, runs, timeout):
    print(f"== BLOCK B: {scn['node_id']} — different question framings, {runs} runs each (match->extract->execute) ==")
    framings = FRAMINGS.get(scn["node_id"]) or [scn["question"] or ""] * 3
    detail_rows, summary_rows = [], []
    all_matched_nodes = set()
    for fi, framing in enumerate(framings, 1):
        matched_nodes, scores, param_hashes, exec_hashes = set(), [], set(), set()
        note_err = None
        for ri in range(1, runs + 1):
            m = _match(framing)
            node = m.get("node_id") if m else None
            score = m.get("score") if m else None
            matched_nodes.add(node)
            if score is not None:
                scores.append(score)
            params, ehash = None, None
            if node == scn["node_id"]:
                try:
                    params = _extract(framing, scn)
                    param_hashes.add(_hash(params.get("resolved", params)))
                    res, e = _run(scn["node_id"], params["filter"], params["group_by"], timeout)
                    ehash = _hash(res) if not e else f"ERR:{e[:40]}"
                    exec_hashes.add(ehash)
                except Exception as exc:  # noqa: BLE001
                    note_err = str(exc)[:120]
            detail_rows.append({
                "framing_#": fi, "framing": framing, "run": ri,
                "matched_node": node or "(no match)",
                "score": round(score, 4) if score is not None else "",
                "params_hash": _hash(params.get("resolved", params)) if params else "",
                "exec_hash": ehash or "",
            })
        all_matched_nodes.update(matched_nodes)
        right_node = matched_nodes == {scn["node_id"]}
        match_stable = len(matched_nodes) == 1
        params_stable = len(param_hashes) <= 1
        exec_stable = len(exec_hashes) <= 1
        verdict = "PASS" if (right_node and params_stable and exec_stable and not note_err) else "FAIL"
        summary_rows.append({
            "framing_#": fi, "framing": framing,
            "matched_node": next(iter(matched_nodes)) if match_stable else f"MIXED:{sorted(str(x) for x in matched_nodes)}",
            "score_range": f"{min(scores):.4f}-{max(scores):.4f}" if scores else "",
            "routes_to_target": "YES" if right_node else "NO",
            "match_stable": "YES" if match_stable else "NO",
            "params_stable": "YES" if params_stable else "NO",
            "exec_deterministic": "YES" if exec_stable else "NO",
            "verdict": verdict, "note": note_err or "",
        })
        print(f"  [{verdict}] framing#{fi}: node={summary_rows[-1]['matched_node']} "
              f"score={summary_rows[-1]['score_range']} params_stable={summary_rows[-1]['params_stable']} "
              f"exec_det={summary_rows[-1]['exec_deterministic']}")
    cross = (all_matched_nodes == {scn["node_id"]})
    print(f"  -> all framings route to a single target node: {'YES' if cross else 'NO ('+str(sorted(str(x) for x in all_matched_nodes))+')'}\n")
    return detail_rows, summary_rows, cross


# --------------------------------------------------------------------------- output


def write_output(out_path, scn, a_rows, a_distinct, a_total, b_detail, b_summary, cross):
    try:
        import openpyxl
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        def sheet(title, headers, rows):
            ws = wb.create_sheet(title)
            ws.append(headers)
            for r in rows:
                ws.append([r.get(h) for h in headers])

        sheet("A_ParamVariation", ["param_set", "group_by", "filter", "runs", "deterministic", "result_hash", "metric", "error"], a_rows)
        sheet("B_Framings_Detail", ["framing_#", "framing", "run", "matched_node", "score", "params_hash", "exec_hash"], b_detail)
        sheet("B_Framings_Summary", ["framing_#", "framing", "matched_node", "score_range", "routes_to_target", "match_stable", "params_stable", "exec_deterministic", "verdict", "note"], b_summary)

        ws = wb.create_sheet("Summary")
        ws.append(["Scenario", scn["node_id"], scn["label"]])
        ws.append([])
        ws.append(["BLOCK A — same scenario, different params"])
        ws.append(["  param sets", a_total])
        ws.append(["  deterministic (4 runs each) PASS", sum(1 for r in a_rows if r["deterministic"] == "PASS"), "of", len(a_rows)])
        ws.append(["  distinct results (params bite)", a_distinct, "of", a_total])
        ws.append([])
        ws.append(["BLOCK B — same scenario, different framings"])
        ws.append(["  framings", len(b_summary)])
        ws.append(["  framings PASS (right node + stable params + deterministic exec)", sum(1 for r in b_summary if r["verdict"] == "PASS"), "of", len(b_summary)])
        ws.append(["  all framings route to target node", "YES" if cross else "NO"])
        wb.save(out_path)
        return out_path
    except ImportError:
        import csv
        base = os.path.splitext(out_path)[0]
        for name, headers, rows in [
            ("A_param", ["param_set", "group_by", "filter", "runs", "deterministic", "result_hash", "metric", "error"], a_rows),
            ("B_detail", ["framing_#", "framing", "run", "matched_node", "score", "params_hash", "exec_hash"], b_detail),
            ("B_summary", ["framing_#", "framing", "matched_node", "score_range", "routes_to_target", "match_stable", "params_stable", "exec_deterministic", "verdict", "note"], b_summary),
        ]:
            with open(f"{base}_{name}.csv", "w", newline="") as fh:
                w = csv.DictWriter(fh, fieldnames=headers)
                w.writeheader()
                w.writerows(rows)
        return f"{base}_*.csv"


def main():
    ap = argparse.ArgumentParser(description="Deep single-scenario consistency test (params + framings, 4x each).")
    ap.add_argument("--scenario", default=DEFAULT_SCENARIO)
    ap.add_argument("--runs", type=int, default=4)
    ap.add_argument("--timeout", type=int, default=240)
    ap.add_argument("--skip-a", action="store_true", help="skip Block A (param variation)")
    ap.add_argument("--skip-b", action="store_true", help="skip Block B (framings; makes LLM calls)")
    ap.add_argument("--out", default=None)
    a = ap.parse_args()

    scn = _fetch_scenario(a.scenario)
    out_path = a.out or f"{a.scenario}_consistency.xlsx"
    print(f"Scenario: {scn['node_id']} — {scn['label']}\n")

    a_rows, a_distinct, a_total = ([], 0, 0)
    if not a.skip_a:
        a_rows, a_distinct, a_total = block_a(scn, a.runs, a.timeout)
    b_detail, b_summary, cross = ([], [], True)
    if not a.skip_b:
        b_detail, b_summary, cross = block_b(scn, a.runs, a.timeout)

    path = write_output(out_path, scn, a_rows, a_distinct, a_total, b_detail, b_summary, cross)
    print("== SUMMARY ==")
    if a_rows:
        print(f"Block A: {sum(1 for r in a_rows if r['deterministic']=='PASS')}/{len(a_rows)} param sets deterministic; "
              f"{a_distinct}/{a_total} distinct results.")
    if b_summary:
        print(f"Block B: {sum(1 for r in b_summary if r['verdict']=='PASS')}/{len(b_summary)} framings PASS; "
              f"all route to target node: {'YES' if cross else 'NO'}.")
    print(f"Wrote: {path}")


if __name__ == "__main__":
    main()
